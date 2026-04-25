#!/usr/bin/env python3
"""
Comparador de precios con fuzzy matching híbrido.
Compara productos propios (~435 de WooCommerce) vs productos de competidores (923 de 3 páginas).

Uso:
    python comparador.py                                    # usa archivos por defecto
    python comparador.py --own mis_productos.csv            # archivo propio custom
    python comparador.py --own X.csv --competitors Y.csv    # ambos archivos custom
    python comparador.py --help                             # ver todas las opciones
"""

import argparse
import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from rapidfuzz import fuzz, process

DEFAULT_OWN = Path("/home/creep/Downloads/wc-product-export-25-4-2026-1777140983496.csv")
DEFAULT_COMPETITORS = Path("resultados_competidores.csv")
DEFAULT_CACHE_DIR = Path("cache/scraping")
DEFAULT_THRESHOLD = 75

BRANDS = {
    "greenpoint", "newmax", "must", "growatt", "deye", "solis", "luxen",
    "trinasolar", "jinko", "ja solar", "apsystems", "epever", "srne", "hoymiles",
    "sunten", "axpert", "mppsolar", "稳固",
}

NOISE_WORDS_OWN = {"(agotado)", "agostado"}
NOISE_WORDS_COMPETITOR = {
    "sin garantia", "sin garantía", "retie", "suntree", "epever",
    "leader", "generico", "genérico", "allyce", "ref", "sku",
}

PATTERNS = {
    "voltage": re.compile(r"(\d+)\s*[vV]"),
    "capacity": re.compile(r"(\d+)\s*[aA][hH]"),
    "power": re.compile(r"(\d+)\s*(?:w|kW|kva|kva)", re.I),
    "brand": re.compile(r"|".join(re.escape(b) for b in BRANDS), re.I),
    "type": re.compile(r"gel|litio|dzf|monocristalino|bifacial|mppt|on-grid|off-grid|hibrido|hibrida|inversor|panel|regulador|batería|bateria", re.I),
}


def normalize_text(text: str) -> str:
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.lower()
    for noise in NOISE_WORDS_OWN | NOISE_WORDS_COMPETITOR:
        text = text.replace(noise, "")
    text = re.sub(r"[^\w\s]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def extract_features(name: str) -> dict:
    name_lower = name.lower()
    features = {}
    if m := PATTERNS["voltage"].search(name):
        features["voltage"] = m.group(1) + "v"
    if m := PATTERNS["capacity"].search(name):
        features["capacity"] = m.group(1) + "ah"
    if m := PATTERNS["power"].search(name):
        features["power"] = m.group(1).lower()
    if m := PATTERNS["brand"].search(name_lower):
        features["brand"] = m.group(0).lower()
    if m := PATTERNS["type"].search(name_lower):
        features["type"] = m.group(0).lower()
    return features


def feature_score(f1: dict, f2: dict) -> int:
    if not f1 or not f2:
        return 0
    matches = sum(1 for k in f1 if k in f2 and f1[k] == f2[k])
    return matches


def normalize_price(price_str: str) -> int:
    if pd.isna(price_str):
        return 0
    s = str(price_str).replace("$", "").replace(" ", "").replace(".", "").replace(",", "")
    try:
        return int(s)
    except ValueError:
        return 0


def load_own_products(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    df.columns = df.columns.str.strip()
    if "Nombre" in df.columns and "precio" not in df.columns:
        df = df.rename(columns={"Nombre": "nombre", "Precio normal": "precio"})
    elif "nombre" in df.columns and "precio" not in df.columns:
        price_col = [c for c in df.columns if "precio" in c.lower()]
        if price_col:
            df = df.rename(columns={price_col[0]: "precio"})
    df["precio"] = df["precio"].apply(normalize_price)
    df["nombre_normalized"] = df["nombre"].apply(normalize_text)
    df["features"] = df["nombre"].apply(extract_features)
    return df


def load_competitor_products(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path)
    df.columns = df.columns.str.strip()
    if "nombre_producto" not in df.columns:
        name_col = [c for c in df.columns if "nombre" in c.lower() or "producto" in c.lower()]
        if name_col:
            df = df.rename(columns={name_col[0]: "nombre_producto"})
    if "precio" in df.columns and "precio_num" not in df.columns:
        df["precio_num"] = df["precio"].apply(normalize_price)
    df["nombre_normalized"] = df["nombre_producto"].apply(normalize_text)
    df["features"] = df["nombre_producto"].apply(extract_features)
    return df


def load_competitor_products_from_cache(cache_dir: Path = DEFAULT_CACHE_DIR) -> pd.DataFrame:
    cached_products = []
    if not cache_dir.exists():
        return pd.DataFrame()

    for item in cache_dir.iterdir():
        if item.is_dir():
            products_file = item / "products.csv"
            if products_file.exists():
                meta_file = item / "meta.json"
                df = pd.read_csv(products_file)
                cached_products.append(df)

    if cached_products:
        combined = pd.concat(cached_products, ignore_index=True)
        combined.columns = combined.columns.str.strip()
        if "precio" in combined.columns and "precio_num" not in combined.columns:
            combined["precio_num"] = combined["precio"].apply(normalize_price)
        combined["nombre_normalized"] = combined["nombre_producto"].apply(normalize_text)
        combined["features"] = combined["nombre_producto"].apply(extract_features)
        return combined
    return pd.DataFrame()


def find_best_match(
    own_row: pd.Series,
    competitors: pd.DataFrame,
    page_name: str,
    threshold: int,
) -> tuple:
    page_df = competitors[competitors["nombre_pagina"] == page_name]
    if page_df.empty:
        return "", 0, 0

    own_features = own_row["features"]
    candidates = page_df.to_dict("records")

    filtered = [
        c for c in candidates
        if feature_score(own_features, c["features"]) >= 2
    ]

    if not filtered:
        filtered = candidates

    choices = [c["nombre_producto"] for c in filtered]
    normalized_choices = [c["nombre_normalized"] for c in filtered]

    if not choices:
        return "", 0, 0

    query = own_row["nombre_normalized"]
    result = process.extractOne(
        query,
        choices,
        scorer=fuzz.token_set_ratio,
        score_cutoff=threshold,
    )

    if result is None:
        return "", 0, 0

    matched_name, score = result[0], result[1]
    matched_record = filtered[choices.index(matched_name)]
    matched_price = matched_record["precio_num"]

    return matched_name, matched_price, int(score)


def diff_percent(own_price: int, comp_price: int) -> float:
    if own_price == 0 or comp_price == 0:
        return 0.0
    return round(((comp_price - own_price) / own_price) * 100, 2)


def generate_multi_sheet_report(report: pd.DataFrame, output_path: Path):
    competitors_pages = ["emergente.com.co", "ingesolar.com.co", "ineldec.com"]

    wb = Workbook()

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    wb.remove(wb.active)

    ws_summary = wb.create_sheet("Resumen")
    headers_summary = ["id", "precio_propio"]
    for page in competitors_pages:
        headers_summary.append(f"precio_{page.replace('.', '_')}")
    ws_summary.append(headers_summary)

    for col in ws_summary.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

    for idx, row in report.iterrows():
        ws_summary.append([idx + 1, row["precio_propio"]] + [
            row.get(f"precio_{page.replace('.', '_')}", 0) for page in competitors_pages
        ])

    for col in ws_summary.columns:
        for cell in col:
            cell.border = border
            cell.alignment = center

    for col in ws_summary.columns:
        ws_summary.column_dimensions[col[0].column_letter].width = 20

    ws_analysis = wb.create_sheet("Analisis Porcentual")
    headers_analysis = ["id", "nombre_producto"]
    for page in competitors_pages:
        headers_analysis.append(f"diff_pct_{page.replace('.', '_')}")
    ws_analysis.append(headers_analysis)

    for col in ws_analysis.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

    for idx, row in report.iterrows():
        ws_analysis.append([idx + 1, row["nombre_producto"]] + [
            row.get(f"diff_pct_{page.replace('.', '_')}", 0) for page in competitors_pages
        ])

    diff_cols = ["C", "D", "E"]
    for col_letter in diff_cols:
        col_idx = ord(col_letter) - ord("A") + 1
        for cell in ws_analysis[col_letter][1:]:
            cell.border = border
            cell.alignment = center
            if cell.value is not None and cell.value != 0:
                if cell.value > 0:
                    cell.fill = red_fill
                elif cell.value < 0:
                    cell.fill = green_fill

    for cell in ws_analysis["A"]:
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for cell in ws_analysis["B"]:
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_analysis.column_dimensions["A"].width = 8
    ws_analysis.column_dimensions["B"].width = 50
    for col_letter in diff_cols:
        ws_analysis.column_dimensions[col_letter].width = 18

    ws_debug = wb.create_sheet("Debug")
    headers_debug = ["id", "nombre_producto"]
    for page in competitors_pages:
        headers_debug.append(f"match_{page.replace('.', '_')}")
        headers_debug.append(f"score_{page.replace('.', '_')}")
    ws_debug.append(headers_debug)

    for col in ws_debug.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

    for idx, row in report.iterrows():
        ws_debug.append([idx + 1, row["nombre_producto"]] + sum([
            [row.get(f"match_{page.replace('.', '_')}", ""),
             row.get(f"score_{page.replace('.', '_')}", 0)] for page in competitors_pages
        ], []))

    ws_debug.column_dimensions["A"].width = 8
    ws_debug.column_dimensions["B"].width = 50
    remaining = 15
    for i, col_letter in enumerate(["C", "E", "G"]):
        ws_debug.column_dimensions[col_letter].width = remaining

    for col in ws_debug.iter_cols(min_row=2):
        for cell in col:
            cell.border = border
            cell.alignment = center

    for cell in ws_debug["B"]:
        cell.alignment = Alignment(horizontal="left", vertical="center")

    wb.save(output_path)


def generate_report(own_df: pd.DataFrame, comp_df: pd.DataFrame, threshold: int) -> pd.DataFrame:
    competitors_pages = ["emergente.com.co", "ingesolar.com.co", "ineldec.com"]

    rows = []
    for _, own_row in own_df.iterrows():
        row = {
            "nombre_producto": own_row["nombre"],
            "precio_propio": own_row["precio"],
        }

        for page in competitors_pages:
            match_name, match_price, score = find_best_match(own_row, comp_df, page, threshold)

            row[f"match_{page.replace('.', '_')}" ] = match_name
            row[f"precio_{page.replace('.', '_')}"] = match_price
            row[f"score_{page.replace('.', '_')}"] = score
            row[f"diff_pct_{page.replace('.', '_')}"] = diff_percent(own_row["precio"], match_price)

        rows.append(row)

    return pd.DataFrame(rows)


def main():
    parser = argparse.ArgumentParser(
        description="Comparador de precios con fuzzy matching híbrido.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python comparador.py                                    # archivos por defecto
  python comparador.py --own mis_productos.csv            # archivo propio custom
  python comparador.py --own X.csv --competitors Y.csv    # ambos archivos custom
  python comparador.py --threshold 80                     # umbral más estricto
  python comparador.py --use-cache                         # usar datos cacheados
  python comparador.py --use-cache --cache-dir cache/scraping  # caché custom
        """
    )
    parser.add_argument("--own", type=Path, default=DEFAULT_OWN,
                        help=f"Ruta al CSV de productos propios (default: {DEFAULT_OWN})")
    parser.add_argument("--competitors", type=Path, default=DEFAULT_COMPETITORS,
                        help=f"Ruta al CSV de competidores (default: {DEFAULT_COMPETITORS})")
    parser.add_argument("--threshold", type=int, default=DEFAULT_THRESHOLD,
                        help=f"Umbral de similitud mínimo 0-100 (default: {DEFAULT_THRESHOLD})")
    parser.add_argument("--output-dir", type=Path, default=Path("."),
                        help="Directorio para archivos de salida (default: ./)")
    parser.add_argument("--use-cache", action="store_true",
                        help="Usar datos de competidores desde el caché")
    parser.add_argument("--cache-dir", type=Path, default=DEFAULT_CACHE_DIR,
                        help=f"Directorio de caché (default: {DEFAULT_CACHE_DIR})")

    args = parser.parse_args()

    if not args.own.exists():
        print(f"Error: archivo propio no encontrado: {args.own}")
        return 1

    SIMILARITY_THRESHOLD = args.threshold

    print("Cargando productos propios...")
    own_df = load_own_products(args.own)
    print(f"  {len(own_df)} productos cargados desde {args.own.name}")

    print("Cargando productos de competidores...")
    if args.use_cache:
        comp_df = load_competitor_products_from_cache(args.cache_dir)
        if comp_df.empty:
            print(f"Error: no se encontraron datos en caché en {args.cache_dir}")
            return 1
        print(f"  {len(comp_df)} productos cargados desde caché")
    elif args.competitors.exists():
        comp_df = load_competitor_products(args.competitors)
        print(f"  {len(comp_df)} productos cargados desde {args.competitors.name}")
    else:
        print(f"Error: archivo de competidores no encontrado: {args.competitors}")
        return 1
    for page in comp_df["nombre_pagina"].unique():
        count = len(comp_df[comp_df["nombre_pagina"] == page])
        print(f"    - {page}: {count}")

    print(f"\nGenerando reporte comparativo (umbral: {SIMILARITY_THRESHOLD})...")
    report = generate_report(own_df, comp_df, SIMILARITY_THRESHOLD)

    output_xlsx = args.output_dir / "reporte_comparativo_precios.xlsx"
    output_csv = args.output_dir / "reporte_comparativo_precios.csv"

    generate_multi_sheet_report(report, output_xlsx)
    print(f"  Guardado: {output_xlsx}")

    report.to_csv(output_csv, index=False)
    print(f"  Guardado: {output_csv}")

    low_confidence = report[
        (report["score_emergente_com_co"] < SIMILARITY_THRESHOLD) &
        (report["score_ingesolar_com_co"] < SIMILARITY_THRESHOLD) &
        (report["score_ineldec_com"] < SIMILARITY_THRESHOLD)
    ]
    if not low_confidence.empty:
        output_revisar = args.output_dir / "matches_revisar.xlsx"
        low_confidence.to_excel(output_revisar, index=False)
        print(f"  Guardado: {output_revisar} ({len(low_confidence)} productos sin match)")

    stats = {
        "total_productos": len(report),
        "matches_emergente": (report["score_emergente_com_co"] >= SIMILARITY_THRESHOLD).sum(),
        "matches_ingesolar": (report["score_ingesolar_com_co"] >= SIMILARITY_THRESHOLD).sum(),
        "matches_ineldec": (report["score_ineldec_com"] >= SIMILARITY_THRESHOLD).sum(),
    }
    print("\nEstadísticas:")
    for k, v in stats.items():
        print(f"  {k}: {v}")

    return 0


if __name__ == "__main__":
    exit(main())
